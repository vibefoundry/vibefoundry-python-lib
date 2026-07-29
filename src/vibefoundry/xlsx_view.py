"""
Render a spreadsheet the way it looks, without a rendering engine.

The approach Google Sheets takes with an uploaded .xlsx: don't reproduce the
file, *reconstruct* it. Cells become an HTML table carrying their real fills,
fonts, borders, merges and number formats; charts are read out of the workbook
as definitions and redrawn. The result reads like the spreadsheet without
needing LibreOffice, Excel, or anything outside pip.

Charts are the part people assume is impossible. They aren't: an .xlsx is a zip,
and xl/charts/chartN.xml holds the chart type, its series, and the cell ranges
each series plots. Files written by Excel also cache the plotted values inline,
so the numbers are usually right there. What we cannot do is match Excel's exact
styling — same as Google Sheets, which draws Google-looking charts from the same
definitions.
"""

import re
import zipfile
from pathlib import Path
from typing import Optional

# Chart kinds we can redraw, mapped to what the frontend should draw. Anything
# else is reported by name rather than silently dropped — a missing chart with
# no explanation is worse than a labelled gap.
CHART_KINDS = {
    "barChart": "bar",
    "bar3DChart": "bar",
    "lineChart": "line",
    "line3DChart": "line",
    "pieChart": "pie",
    "pie3DChart": "pie",
    "doughnutChart": "doughnut",
    "scatterChart": "scatter",
    "areaChart": "area",
    "area3DChart": "area",
    "radarChart": "radar",
}


def _local(tag: str) -> str:
    """Strip any XML namespace. Chart XML is written with a 'c:' prefix by Excel
    and as the default namespace by openpyxl, so matching on the bare local name
    is the only thing that works for both."""
    return tag.rsplit("}", 1)[-1].rsplit(":", 1)[-1]


def _iter_local(node, name: str):
    for child in node.iter():
        if _local(child.tag) == name:
            yield child


def _first_local(node, name: str):
    return next(_iter_local(node, name), None)


def _text_of(node) -> str:
    """Concatenate every <a:t> run — chart titles are split across runs."""
    if node is None:
        return ""
    return "".join(t.text or "" for t in _iter_local(node, "t")).strip()


def _cached(ref_parent) -> tuple[list, bool]:
    """Values cached inside the chart, in index order.

    Excel writes numCache/strCache so a chart renders without recalculating.
    That means we usually get the plotted numbers without touching the sheet —
    and get them *as plotted*, which is what we want.
    """
    pts = {}
    for cache_name in ("numCache", "strCache"):
        cache = _first_local(ref_parent, cache_name)
        if cache is None:
            continue
        for pt in _iter_local(cache, "pt"):
            idx = pt.get("idx")
            v = _first_local(pt, "v")
            if idx is None or v is None:
                continue
            raw = (v.text or "").strip()
            if cache_name == "numCache":
                try:
                    raw = float(raw)
                except ValueError:
                    raw = None
            pts[int(idx)] = raw
    if not pts:
        return [], False
    return [pts.get(i) for i in range(max(pts) + 1)], True


def _ref_formula(ref_parent) -> Optional[str]:
    f = _first_local(ref_parent, "f")
    return (f.text or "").strip() if f is not None and f.text else None


_RANGE = re.compile(r"^(?:'([^']+)'|([^!]+))!\$?([A-Z]+)\$?(\d+)(?::\$?([A-Z]+)\$?(\d+))?$")


def _resolve(workbook, formula: Optional[str]) -> list:
    """Read a range like 'Sales'!$B$2:$B$5 out of the workbook.

    The fallback for charts with no cached values (openpyxl-written files have
    none). Returns [] rather than raising for anything unparseable — a chart
    with missing data should degrade, not break the whole preview.
    """
    if not formula or workbook is None:
        return []
    m = _RANGE.match(formula.strip())
    if not m:
        return []
    sheet_name = m.group(1) or m.group(2)
    if sheet_name not in workbook.sheetnames:
        return []
    ws = workbook[sheet_name]
    c1, r1 = m.group(3), int(m.group(4))
    c2, r2 = m.group(5) or c1, int(m.group(6) or r1)
    out = []
    try:
        for row in ws[f"{c1}{r1}:{c2}{r2}"]:
            for cell in row:
                out.append(cell.value)
    except (ValueError, KeyError):
        return []
    return out


def _series_from(ser, workbook) -> dict:
    """One <ser> element -> {name, values}."""
    tx = _first_local(ser, "tx")
    name = ""
    if tx is not None:
        # Three ways a series name is stored, all seen in the wild:
        #   <tx><strRef><strCache><pt><v>   cached (Excel)
        #   <tx><strRef><f>                 a cell reference (openpyxl)
        #   <tx><v>                         the name inline (xlsxwriter)
        cached, ok = _cached(tx)
        name = str(cached[0]) if ok and cached and cached[0] is not None else ""
        if not name:
            f = _ref_formula(tx)
            resolved = _resolve(workbook, f)
            name = str(resolved[0]) if resolved and resolved[0] is not None else ""
        if not name:
            # A <v> sitting directly under <tx>, rather than inside a cache.
            direct = next((v for v in tx if _local(v.tag) == "v"), None)
            if direct is not None and direct.text:
                name = direct.text.strip()
        if not name:
            name = _ref_formula(tx) or ""

    # scatter uses yVal; everything else uses val
    holder = _first_local(ser, "val") or _first_local(ser, "yVal")
    values: list = []
    if holder is not None:
        values, ok = _cached(holder)
        if not ok:
            values = _resolve(workbook, _ref_formula(holder))
    return {
        "name": name,
        "values": [v if isinstance(v, (int, float)) or v is None else _as_number(v) for v in values],
    }


def _as_number(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _categories_from(ser, workbook) -> list:
    holder = _first_local(ser, "cat") or _first_local(ser, "xVal")
    if holder is None:
        return []
    cats, ok = _cached(holder)
    if not ok:
        cats = _resolve(workbook, _ref_formula(holder))
    return ["" if c is None else str(c) for c in cats]


def _charts_for_sheet(zf, sheet_name: Optional[str]) -> Optional[set]:
    """Chart parts belonging to one sheet, or None if the chain can't be walked.

    A workbook's charts are reachable only through relationships:
    workbook -> sheetN.xml -> drawingM.xml -> chartK.xml. Without following it,
    every sheet shows every chart in the file, which is what a 5-sheet workbook
    made obvious. Returning None means "couldn't tell" and the caller falls back
    to showing all of them — better than showing none.
    """
    import xml.etree.ElementTree as ET
    import posixpath

    if not sheet_name:
        return None

    def rels_for(part: str) -> dict:
        rel_path = posixpath.join(posixpath.dirname(part), "_rels", posixpath.basename(part) + ".rels")
        try:
            root = ET.fromstring(zf.read(rel_path))
        except (KeyError, ET.ParseError):
            return {}
        out = {}
        for rel in root:
            rid, target = rel.get("Id"), rel.get("Target")
            if not rid or not target:
                continue
            # A Target is normally relative to the part's directory and may climb
            # out ("../drawings/x.xml"), but writers also emit it package-absolute
            # ("/xl/worksheets/sheet1.xml"). Zip entries carry no leading slash,
            # so an absolute target has to be stripped or every lookup misses.
            if target.startswith("/"):
                resolved = target.lstrip("/")
            else:
                resolved = posixpath.normpath(posixpath.join(posixpath.dirname(part), target))
            out[rid] = resolved
        return out

    try:
        wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
    except (KeyError, ET.ParseError):
        return None

    wb_rels = rels_for("xl/workbook.xml")
    sheet_part = None
    for sheet_el in _iter_local(wb_root, "sheet"):
        if sheet_el.get("name") != sheet_name:
            continue
        rid = next((v for k, v in sheet_el.attrib.items() if k.endswith("}id") or k == "r:id"), None)
        sheet_part = wb_rels.get(rid)
        break
    if not sheet_part:
        return None

    charts = set()
    for drawing in rels_for(sheet_part).values():
        if "drawing" not in drawing:
            continue
        for target in rels_for(drawing).values():
            if re.search(r"chart\d+\.xml$", target):
                charts.add(target)
    return charts


def extract_charts(path: Path, workbook=None, sheet: Optional[str] = None) -> list[dict]:
    """Every chart in the workbook, as a spec the frontend can draw.

    Never raises: a malformed chart part yields fewer charts, not a failed
    preview.
    """
    import xml.etree.ElementTree as ET

    charts: list[dict] = []
    try:
        zf = zipfile.ZipFile(path)
    except (zipfile.BadZipFile, OSError):
        return charts

    with zf:
        # Writers disagree about where chart parts live: openpyxl and Excel use
        # xl/charts/chartN.xml, xlsxwriter uses xl/drawings/charts/chartN.xml.
        # Match either — looking in only one place silently loses every chart in
        # files from the other, which is exactly how this was first found.
        parts = sorted(
            n for n in zf.namelist()
            if re.match(r"^xl/(?:\w+/)*charts?/chart\d+\.xml$", n)
        )
        # Restrict to the sheet being shown. None means the relationship chain
        # could not be walked, in which case show everything rather than nothing.
        owned = _charts_for_sheet(zf, sheet)
        if owned is not None:
            scoped = [n for n in parts if n in owned]
            # An empty result is meaningful (this sheet has no charts); only
            # ignore the filter if it found nothing anywhere, which would mean
            # the paths didn't line up.
            if scoped or not owned:
                parts = scoped
        for part in parts:
            try:
                root = ET.fromstring(zf.read(part))
            except (ET.ParseError, KeyError):
                continue

            plot = _first_local(root, "plotArea")
            if plot is None:
                continue

            kind_el = next(
                (c for c in plot if _local(c.tag) in CHART_KINDS or _local(c.tag).endswith("Chart")),
                None,
            )
            if kind_el is None:
                continue
            kind_name = _local(kind_el.tag)
            kind = CHART_KINDS.get(kind_name)

            title = _text_of(_first_local(root, "title"))

            if kind is None:
                # Named, not dropped — the user can see what didn't come through.
                charts.append({"unsupported": kind_name, "title": title})
                continue

            # Column vs bar is an attribute, not a separate chart type.
            if kind == "bar":
                bar_dir = _first_local(kind_el, "barDir")
                if bar_dir is not None and bar_dir.get("val") == "bar":
                    kind = "horizontalBar"

            sers = list(_iter_local(kind_el, "ser"))
            if not sers:
                continue

            categories = _categories_from(sers[0], workbook)
            series = [_series_from(s, workbook) for s in sers]
            series = [s for s in series if any(v is not None for v in s["values"])]
            if not series:
                continue

            charts.append({"type": kind, "title": title, "categories": categories, "series": series})

    return charts


_LITERAL_CURRENCY = re.compile(r'"([$£€¥])"')
_WHITE_TEXT = re.compile(r'style="([^"]*)"')
_STYLE_ATTR = re.compile(r'\sstyle="([^"]*)"')

# Beyond this a sheet is shown truncated. The rendered view is for reading a
# spreadsheet, not scrolling a dataset — that is what the data view is for, and
# it virtualizes. Without a cap a 100k-row sheet builds a 100k-row DOM.
MAX_RENDERED_ROWS = 2000


_ROW_OPEN = re.compile(r"<tr\b", re.IGNORECASE)


def _truncate_rows(html: str, limit: int = MAX_RENDERED_ROWS) -> tuple[str, int]:
    """Keep at most `limit` rows. Returns (html, rows_dropped).

    Cuts at a row boundary and closes the table, so the markup stays valid.
    """
    starts = [m.start() for m in _ROW_OPEN.finditer(html)]
    if len(starts) <= limit:
        return html, 0
    cut = starts[limit]
    tail = html[cut:]
    # Preserve whatever closed the table after the last kept row.
    closing = "".join(t for t in ("</tbody>", "</table>", "</body>", "</html>") if t in tail.lower())
    return html[:cut] + closing, len(starts) - limit


def _dedupe_styles(html: str) -> str:
    """Collapse repeated inline styles into classes.

    Every cell arrives carrying its full style inline, and a sheet's cells share
    a handful of styles between them — so a 400-row sheet spends most of its
    ~280KB repeating the same few declarations. Hoisting them into classes cuts
    the payload several-fold, which matters most in a pane, where every byte
    crosses the host bridge on each sheet switch.
    """
    seen: dict[str, str] = {}

    def swap(m: "re.Match") -> str:
        style = m.group(1)
        if not style:
            return ""
        cls = seen.get(style)
        if cls is None:
            cls = f"c{len(seen)}"
            seen[style] = cls
        return f' class="{cls}"'

    body = _STYLE_ATTR.sub(swap, html)
    if not seen:
        return html
    rules = "".join(f".{cls}{{{style}}}" for style, cls in seen.items())
    return f"<style>{rules}</style>{body}"


def _clean_html(html: str) -> str:
    """Repair two things the cell renderer gets wrong on real workbooks.

    1. Excel number formats quote literal text — `"$"#,##0` means a dollar sign
       followed by a number. Those quotes are delimiters, not characters, but
       they arrive rendered, so currency reads as `"$"42,000`.

    2. A header styled white-on-a-theme-fill comes through as white text with no
       background, because theme colours aren't resolved — leaving the row
       invisible against the page. Rather than guess the intended fill, darken
       text that would otherwise be unreadable. Losing the intended colour is a
       far smaller failure than losing the words.
    """
    html = _LITERAL_CURRENCY.sub(r"\1", html)

    def fix_style(m: "re.Match") -> str:
        style = m.group(1)
        if "background-color" in style:
            return m.group(0)  # it has a fill; white text is legible on it
        # #fff / #ffffff / #FFFFFFFF, with or without an alpha pair
        if re.search(r"color:\s*#(?:fff|ffff|ffffff|ffffffff)\b", style, re.IGNORECASE):
            style = re.sub(
                r"color:\s*#(?:fff|ffff|ffffff|ffffffff)\b",
                "color: #0d0d0d",
                style,
                flags=re.IGNORECASE,
            )
        return f'style="{style}"'

    return _dedupe_styles(_WHITE_TEXT.sub(fix_style, html))


def sheet_names(path: Path) -> list[str]:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names
    except Exception:
        return []


# Renders keyed by (path, mtime, size, sheet). Switching between sheets of one
# workbook is the common case and re-parses the whole file each time otherwise.
_CACHE: "dict[tuple, dict]" = {}
_CACHE_LIMIT = 24


def _cache_key(path: Path, sheet: Optional[str]):
    st = path.stat()
    return (str(path.resolve()), st.st_size, int(st.st_mtime), sheet)


def render(path: Path, sheet: Optional[str] = None) -> Optional[dict]:
    """{html, sheets, activeSheet, charts} — or None if it can't be read."""
    try:
        from xlsx2html import xlsx2html
    except ImportError:
        return None

    try:
        key = _cache_key(path, sheet)
    except OSError:
        key = None
    if key is not None and key in _CACHE:
        return _CACHE[key]

    names = sheet_names(path)
    active = sheet if sheet and sheet in names else (names[0] if names else None)

    try:
        import io

        buf = io.StringIO()
        # append_headers/append_lineno draw Excel's A/B/C and 1/2/3 gutters.
        xlsx2html(str(path), buf, sheet=active)
        html, dropped = _truncate_rows(buf.getvalue())
        truncated = dropped or 0
        html = _clean_html(html)
    except Exception:
        return None

    workbook = None
    try:
        from openpyxl import load_workbook

        # data_only: we want the values Excel computed, not formula strings.
        workbook = load_workbook(path, data_only=True)
    except Exception:
        workbook = None

    try:
        charts = extract_charts(path, workbook, active)
    except Exception:
        charts = []
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass

    result = {
        "html": html,
        "sheets": names,
        "activeSheet": active,
        "charts": charts,
        "truncated": truncated,
    }
    if key is not None:
        if len(_CACHE) >= _CACHE_LIMIT:
            _CACHE.pop(next(iter(_CACHE)))
        _CACHE[key] = result
    return result
